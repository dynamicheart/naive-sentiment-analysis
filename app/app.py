# -*- coding: UTF-8 -*-

import MySQLdb
from openpyxl import Workbook

from sentiment import SentimentAnalyzer

def main():
    db = MySQLdb.connect("localhost", "weibospider", "123456", "weibo", charset='utf8mb4' )
    with db:
        cursor = db.cursor()
        cursor.execute(""" SELECT *
            FROM weibo_data
            WHERE weibo_video != '' """)

        weibos = cursor.fetchall()

        analyzer = SentimentAnalyzer('../data_set/stopwords.txt', '../data_set/notdict.txt',
        '../data_set/BosonNLP_sentiment_score.txt', '../data_set/degreedict.txt')

        wb = Workbook()
        ws = wb.active
        ws.title = "汇总表"

        # Print titles
        ws['A1'] = '时间'
        ws['B1'] = '微博内容'
        ws['C1'] = '微博URL'
        ws['D1'] = '点赞数'
        ws['E1'] = '分享数'
        ws['F1'] = '评论数'
        ws['G1'] = '情感均值'

        weibo_row_idx = 2
        for weibo in weibos:
            cursor.execute("""SELECT *
                FROM weibo_comment
                WHERE weibo_id = %s
                """, (weibo[1],))

            ws['A' + str(weibo_row_idx)] = weibo[13] # create_time
            ws['B' + str(weibo_row_idx)] = weibo[2] # weibo_cont
            ws['C' + str(weibo_row_idx)] = weibo[12] # weibo_url
            ws['D' + str(weibo_row_idx)] = weibo[8] # praise_num
            ws['E' + str(weibo_row_idx)] = weibo[6] # repost_num
            ws['F' + str(weibo_row_idx)] = weibo[7] # comment_num

            comments = cursor.fetchall()

            weibo_ws = wb.create_sheet('微博' + str(weibo_row_idx - 1))
            weibo_ws['A1'] = '短视频微博1(' + weibo[2] + ')'
            weibo_ws.merge_cells(start_row = 1, start_column = 1, \
                            end_row = 1, end_column = 6)

            weibo_ws['B2'] = '评论内容'
            weibo_ws['C2'] = '情感值'

            comment_row_idx = 3
            total = 0
            for comment in comments:
                score = analyzer.calcuate_score(comment[2]) # comment_cont
                weibo_ws['B' + str(comment_row_idx)] = comment[2]
                weibo_ws['C' + str(comment_row_idx)] = score
                total += score
                comment_row_idx += 1

            avg = 0
            if len(comments) > 0:
                avg = total / len(comments)

            print(total, len(comments), avg)
            weibo_ws['A' + str(comment_row_idx)] = '总计'
            weibo_ws['B' + str(comment_row_idx)] = len(comments)
            comment_row_idx += 1
            weibo_ws['A' + str(comment_row_idx)] = '情感均值'
            weibo_ws['C' + str(comment_row_idx)] = avg

            ws['G' + str(weibo_row_idx)] = avg; # 情感均值
            weibo_row_idx += 1

        wb.save('nike_weibo.xlsx')


if __name__ == '__main__':
    main()
